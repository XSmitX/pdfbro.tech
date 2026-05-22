"""PDF to Excel converter using PyMuPDF and openpyxl.

Usage:
  python convert_pdf_to_excel.py <input.pdf> <output.xlsx>

Exit codes:
  0 - success (prints SUCCESS:<output_path>)
  1 - error (message printed to stderr)
"""

import os
import re
import sys


def _safe_sheet_name(name: str, index: int) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name).strip()
    if not cleaned:
        cleaned = f"Sheet{index}"
    return cleaned[:31]


def convert_pdf_to_xlsx(input_path: str, output_path: str) -> None:
    try:
        import fitz  # PyMuPDF
        from openpyxl import Workbook

        doc = fitz.open(input_path)
        if doc.page_count == 0:
            raise RuntimeError("The input PDF has no pages")

        total_pages = doc.page_count

        wb = Workbook()
        wb.remove(wb.active)

        table_count = 0
        for page_index, page in enumerate(doc, start=1):
            ws = wb.create_sheet(title=_safe_sheet_name(f"Page {page_index}", page_index))
            row_ptr = 1

            tables = page.find_tables()
            extracted_tables = [t.extract() for t in tables.tables] if tables.tables else []

            if extracted_tables:
                for t_index, table_rows in enumerate(extracted_tables, start=1):
                    if t_index > 1:
                        row_ptr += 1
                    ws.cell(row=row_ptr, column=1, value=f"Table {t_index}")
                    row_ptr += 1

                    for row in table_rows:
                        values = row if row else []
                        for col_idx, value in enumerate(values, start=1):
                            ws.cell(row=row_ptr, column=col_idx, value="" if value is None else str(value))
                        row_ptr += 1

                    row_ptr += 1
                    table_count += 1
            else:
                # Fallback when no tables are detected: keep page text in a single column.
                text = page.get_text("text")
                lines = [line.strip() for line in text.splitlines() if line.strip()]
                if lines:
                    ws.cell(row=1, column=1, value="No structured table detected on this page.")
                    for idx, line in enumerate(lines, start=3):
                        ws.cell(row=idx, column=1, value=line)
                else:
                    ws.cell(row=1, column=1, value="No extractable text found on this page.")

        wb.save(output_path)
        doc.close()
        print(
            f"Converted {total_pages} page(s) to Excel. Extracted {table_count} table(s).",
            file=sys.stderr,
        )

    except ImportError as e:
        missing = str(e).split("'")[1] if "'" in str(e) else str(e)
        raise RuntimeError(
            f"Required library not installed: {missing}. "
            "Run: pip install pymupdf openpyxl"
        )
    except Exception as e:
        raise RuntimeError(f"Conversion failed: {e}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python convert_pdf_to_excel.py <input.pdf> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    if not os.path.exists(input_file):
        print(f"Input file not found: {input_file}", file=sys.stderr)
        sys.exit(1)

    try:
        convert_pdf_to_xlsx(input_file, output_file)
        if os.path.exists(output_file):
            print(f"SUCCESS:{output_file}")
            sys.exit(0)
        print("Conversion produced no output file", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(str(e), file=sys.stderr)
        sys.exit(1)
